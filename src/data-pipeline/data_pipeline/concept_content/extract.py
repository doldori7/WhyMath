"""통합마스터 xlsx → K-12 개념 콘텐츠 모델 (openpyxl 지연 import + 조인).

`개념` 시트(학교급!='대학교')의 콘텐츠(은유·오개념·정식정의·허용표현·설명)와 `암기카드(목록화)`
시트(개념ID)의 암기카드를 **개념ID로 조인**해 `ConceptContent`를 만든다. 대학(학교급=='대학교')
행은 읽지 않는다(U4 별도). openpyxl은 `[xlsx]` extra라 지연 import.

**redaction**: `성취기준 문장`·`성취기준 요약(간결)` 컬럼(NCIC 본문)은 **모델에 절대 싣지 않는다**.
`연결 성취기준` 컬럼은 `standard_codes`로 **코드만** 파싱(';'·',' 분할, '없음'·빈값 → []).
"""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from typing import Any, Final

from data_pipeline.concept_content.models import ConceptContent, Flashcard

SHEET_CONCEPTS: Final[str] = "개념"
SHEET_CARDS: Final[str] = "암기카드(목록화)"
_UNIVERSITY: Final[str] = "대학교"

# NCIC 본문 컬럼 — 모델 미수록(redaction). 방어적으로 이름을 박아둔다.
REDACTED_COLUMNS: Final[frozenset[str]] = frozenset({"성취기준 문장", "성취기준 요약(간결)"})


class ExtractError(ValueError):
    """xlsx 추출 실패."""


def _load_workbook(path: str | Path) -> Any:
    try:
        import openpyxl
    except ModuleNotFoundError as exc:  # pragma: no cover - 설치 환경 의존
        raise ExtractError("openpyxl 미설치 — `pip install -e '.[xlsx]'` 후 재시도.") from exc
    return openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)


def _iter_rows(workbook: Any, sheet: str) -> Iterator[dict[str, object]]:
    if sheet not in workbook.sheetnames:
        raise ExtractError(f"시트 없음: {sheet!r} (가용: {list(workbook.sheetnames)})")
    row_iter = workbook[sheet].iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration:
        return
    headers = [str(c).strip() if c is not None else "" for c in header_row]
    for raw in row_iter:
        if all(c is None for c in raw):
            continue
        yield {h: (raw[i] if i < len(raw) else None) for i, h in enumerate(headers) if h}


def _opt(row: dict[str, object], key: str) -> str | None:
    """선택 문자열 — None/공란/'-'/'None'은 None."""
    value = row.get(key)
    text = "" if value is None else str(value).strip()
    return None if (not text or text in {"-", "None"}) else text


def _parse_standard_codes(row: dict[str, object]) -> list[str]:
    """`연결 성취기준` → 코드 리스트(NCIC 본문 다리). ';'·',' 분할·'없음'/빈값 → []."""
    value = row.get("연결 성취기준")
    text = "" if value is None else str(value).strip()
    if not text or text in {"없음", "-", "None"}:
        return []
    codes: list[str] = []
    for chunk in text.split(";"):
        for token in chunk.split(","):
            code = token.strip()
            if code and code not in {"없음", "-", "None"}:
                codes.append(code)
    return codes


def _card_from_row(row: dict[str, object]) -> Flashcard:
    return Flashcard(
        front=str(row.get("앞면(front)", "")).strip(),
        back=str(row.get("뒷면(back)", "")).strip(),
        mnemonic=_opt(row, "암기보조(mnemonic)"),
        exposure_condition=_opt(row, "노출조건"),
        grade=_opt(row, "등급"),
        difficulty_tier=_opt(row, "난이도층"),
    )


def extract_k12_content(path: str | Path) -> list[ConceptContent]:
    """K-12 개념 콘텐츠 목록 — `개념`(학교급!='대학교') + `암기카드`(개념ID) 조인.

    카드는 개념ID로 묶어 해당 콘텐츠에 붙인다(개념당 0개 이상). 콘텐츠 순서는 `개념` 시트 순서를
    보존한다(결정론). `성취기준 문장`·`성취기준 요약(간결)`(NCIC 본문)은 모델에 싣지 않으며
    `연결 성취기준`은 코드만 보존한다.
    """
    workbook = _load_workbook(path)
    try:
        concept_rows = [
            r
            for r in _iter_rows(workbook, SHEET_CONCEPTS)
            if str(r.get("학교급", "")).strip() != _UNIVERSITY
        ]
        k12_codes = {str(r.get("개념ID", "")).strip() for r in concept_rows}
        cards_by_code: dict[str, list[Flashcard]] = {}
        for r in _iter_rows(workbook, SHEET_CARDS):
            code = str(r.get("개념ID", "")).strip()
            if code in k12_codes:
                cards_by_code.setdefault(code, []).append(_card_from_row(r))
    finally:
        workbook.close()

    out: list[ConceptContent] = []
    for r in concept_rows:
        code = str(r.get("개념ID", "")).strip()
        out.append(
            ConceptContent(
                code=code,
                name=str(r.get("개념명", "")).strip(),
                subject=str(r.get("과목", "")).strip(),
                unit=_opt(r, "단원(영역)"),
                metaphor=_opt(r, "은유"),
                misconception=_opt(r, "오개념"),
                formal_definition_internal=_opt(r, "정식정의(내부·학생비노출)"),
                accepted_expressions=_opt(r, "허용표현(인정 표현)"),
                explanation=_opt(r, "설명"),
                standard_codes=_parse_standard_codes(r),
                flashcards=cards_by_code.get(code, []),
            )
        )
    return out


__all__ = [
    "REDACTED_COLUMNS",
    "SHEET_CARDS",
    "SHEET_CONCEPTS",
    "ExtractError",
    "extract_k12_content",
]
