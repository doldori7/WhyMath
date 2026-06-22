"""대학 통합마스터 xlsx → 대학 소단원 콘텐츠 모델 (openpyxl 지연 import + 조인).

`개념` 시트(학교급=='대학교')의 콘텐츠(은유·오개념·정식정의·허용표현·설명)와 `암기카드(목록화)`
시트(개념ID=소단원 코드)의 암기카드를 **소단원 코드로 조인**해 `UniversityConceptContent`를 만든다.
대학 외(초중고) 행은 읽지 않는다(라이선스 분리). openpyxl은 `[xlsx]` extra라 지연 import.
"""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from typing import Any, Final

from data_pipeline.concept_content_university.models import (
    SUBUNIT_PATTERN,
    Flashcard,
    UniversityConceptContent,
)

SHEET_CONCEPTS: Final[str] = "개념"
SHEET_CARDS: Final[str] = "암기카드(목록화)"


class ExtractError(ValueError):
    """xlsx 추출 실패."""


def _load_workbook(path: str | Path) -> Any:
    try:
        import openpyxl
    except ModuleNotFoundError as exc:  # pragma: no cover - 설치 환경 의존
        raise ExtractError("openpyxl 미설치 — `pip install -e '.[xlsx]'` 후 재시도.") from exc
    return openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)


def _iter_rows(workbook: Any, sheet: str) -> Iterator[dict[str, object]]:
    if sheet not in workbook.sheetnames:
        raise ExtractError(f"시트 없음: {sheet!r} (가용: {list(workbook.sheetnames)})")
    row_iter = workbook[sheet].iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration:
        return
    headers = [str(c).strip() if c is not None else "" for c in header_row]
    for raw in row_iter:
        if all(c is None for c in raw):
            continue
        yield {h: (raw[i] if i < len(raw) else None) for i, h in enumerate(headers) if h}


def _opt(row: dict[str, object], key: str) -> str | None:
    """선택 문자열 — None/공란/'-'/'None'은 None."""
    value = row.get(key)
    text = "" if value is None else str(value).strip()
    return None if (not text or text in {"-", "None"}) else text


def _card_from_row(row: dict[str, object]) -> Flashcard:
    return Flashcard(
        front=str(row.get("앞면(front)", "")).strip(),
        back=str(row.get("뒷면(back)", "")).strip(),
        mnemonic=_opt(row, "암기보조(mnemonic)"),
        exposure_condition=_opt(row, "노출조건"),
        grade=_opt(row, "등급"),
        difficulty_tier=_opt(row, "난이도층"),
    )


def extract_university_content(path: str | Path) -> list[UniversityConceptContent]:
    """대학 소단원 콘텐츠 목록 — `개념`(대학) + `암기카드`(소단원 코드) 조인.

    카드는 소단원 코드로 묶어 해당 콘텐츠에 붙인다(소단원당 0개 이상). 콘텐츠 순서는 `개념` 시트
    순서를 보존한다(결정론).
    """
    workbook = _load_workbook(path)
    try:
        concept_rows = [
            r
            for r in _iter_rows(workbook, SHEET_CONCEPTS)
            if str(r.get("학교급", "")).strip() == "대학교"
        ]
        cards_by_code: dict[str, list[Flashcard]] = {}
        for r in _iter_rows(workbook, SHEET_CARDS):
            code = str(r.get("개념ID", "")).strip()
            if SUBUNIT_PATTERN.match(code):
                cards_by_code.setdefault(code, []).append(_card_from_row(r))
    finally:
        workbook.close()

    out: list[UniversityConceptContent] = []
    for r in concept_rows:
        code = str(r.get("개념ID", "")).strip()
        out.append(
            UniversityConceptContent(
                code=code,
                name=str(r.get("개념명", "")).strip(),
                subject=str(r.get("과목", "")).strip(),
                unit=_opt(r, "단원(영역)"),
                metaphor=_opt(r, "은유"),
                misconception=_opt(r, "오개념"),
                formal_definition_internal=_opt(r, "정식정의(내부·학생비노출)"),
                accepted_expressions=_opt(r, "허용표현(인정 표현)"),
                explanation=_opt(r, "설명"),
                flashcards=cards_by_code.get(code, []),
            )
        )
    return out


__all__ = [
    "SHEET_CARDS",
    "SHEET_CONCEPTS",
    "ExtractError",
    "extract_university_content",
]
