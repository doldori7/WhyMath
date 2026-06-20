"""File A(성취기준·개념그래프 통합 마스터 xlsx) → 모델 추출기.

P1-1. File A의 시트를 읽어 `AchievementStandard`·`ConceptStandardLink`로 변환한다.

레이어 경계:
  - 순수 행 파서(`parse_standard_row`·`parse_link_row`)는 transform 레이어만 사용한다 →
    openpyxl 없이도 import·호출 가능([postgres]·[neo4j] 드라이버 지연 import 선례와 동형).
  - openpyxl(`[xlsx]` extra)은 `extract_file_a`에서만 **지연 import**한다.

re-ID/스키마 안정성:
  - norm_id는 행에 직접 있으면 사용하고, 없으면 transform이 `build_norm_id(개정, code)`로 파생한다.
  - 연결의 개념 참조는 안정 키 `concept_src_id`(File A 원본키)를 그대로 보존한다 — concept_id
    마이그레이션(UC→ELEM-GEO)에 견고. 현재 개념 identity 해소는 백엔드 로더 책임.

✅ File A 도착(2026-06-20·sha256 `4e21d6d4…`): 아래 `_SHEET_*`·`_STD_COLS`·`_LINK_COLS`
좌변을 실 File A 컬럼명으로 정합 완료. 파서·검증·모델·패턴은 무변경(좌변만 조정). 실 추출 결과:
성취기준 895·연결 437행(다중코드 분할로 443 링크) 클린(`external_corpus_ingestion_v1.md` §1).

실데이터 특성 3종(연결 시트 `개념-성취기준-CCSS`): ① 교육과정 열 없음 → revision 2022 기본,
② `연결구분` 괄호 접미사(`직접(기본수학)`) → 베어 토큰 정규화·괄호는 note 보존, ③ code 셀 `;`
다중코드 → 코드별 1 링크 분할. 표준 시트의 `norm_id` 열은 **로마숫자 Ⅰ 유지**(ASCII 패턴 위반)라
매핑하지 않고 `build_norm_id`(Ⅰ→I 정규화) 파생을 쓴다.
"""

from __future__ import annotations

import re
from collections.abc import Iterator, Mapping
from pathlib import Path
from typing import Any, Final, cast

from data_pipeline.ncic.models import AchievementStandard, ConceptStandardLink
from data_pipeline.ncic.transform import (
    RawLinkRecord,
    RawStandardRecord,
    build_norm_id,
    transform_raw_to_link,
    transform_raw_to_standard,
)

# ──────────────────────────────────────────────────────────────────────
# File A 시트·컬럼 매핑 (실 File A 정합 — 좌변=실제 컬럼명)
# ──────────────────────────────────────────────────────────────────────
# 전체 성취기준(895행·2022+2015). 학교급·학년군은 코드에서 추론하므로 매핑 불요.
_SHEET_STANDARDS: Final[str] = "성취기준_목록"
# 개념↔성취기준 연결(연결구분 = 직접/재매핑/준용). 실 File A 시트명.
_SHEET_LINKS: Final[str] = "개념-성취기준-CCSS"
# 2022 전용(공식 마스터) — official_code↔norm_id 교차검증용. 현재 미배선(후속).
_SHEET_OFFICIAL_MASTER: Final[str] = "공식_성취기준_마스터"

# 성취기준 컬럼: File A 컬럼명 → RawStandardRecord 키.
#   - code·statement는 필수, 그 외 선택(미존재 시 transform 기본값/추론).
#   - norm_id 컬럼은 매핑하지 않는다(File A norm_id는 로마숫자 Ⅰ 유지 → ASCII 패턴 위반).
#     transform이 (교육과정, code)로 Ⅰ→I 정규화하며 결정적 파생한다.
#   - 해설·핵심아이디어·출처URL·출처문서는 실 File A에 열이 없으면 자동 skip(미존재=무해).
_STD_COLS: Final[dict[str, str]] = {
    "성취기준 코드": "code",
    "교육과정": "curriculum_revision",
    "과목약칭": "subject",
    "영역(단원)": "domain",
    "소단원(개념)": "sub_domain",
    "성취기준 내용": "statement",
    "해설": "commentary",
    "핵심아이디어": "big_idea",
    "출처URL": "source_url",
    "출처문서": "source_document",
}

# 연결 컬럼: File A 컬럼명 → 내부 키.
#   - 실 링크 시트엔 code 열만 있고 교육과정·norm_id·비고 열이 없다 → revision 2022 기본·파생.
#   - 하위호환: `norm_id`/`코드`/`교육과정`/`비고` 열도 매핑해 직접 norm_id·구버전 시트를 수용.
_LINK_COLS: Final[dict[str, str]] = {
    "개념ID": "concept_src_id",
    "한국 성취기준(2022)": "code",
    "연결구분": "link_type",
    "norm_id": "norm_id",
    "코드": "code",
    "교육과정": "curriculum_revision",
    "비고": "note",
}

# source_url 의무(공공누리 1유형) — 행에 출처 URL이 없을 때 NCIC 기본 출처로 채운다.
_DEFAULT_SOURCE_URL: Final[str] = "https://www.ncic.go.kr"

# 교육과정 표기 정규화 — '2022개정'(공백 없음) 등 변형을 모델이 요구하는 '2022 개정'으로.
_REVISION_CANON: Final[dict[str, str]] = {
    "2022개정": "2022 개정",
    "2022 개정": "2022 개정",
    "2015개정": "2015 개정",
    "2015 개정": "2015 개정",
}


class ExtractError(ValueError):
    """xlsx 추출 실패."""


# ──────────────────────────────────────────────────────────────────────
# 셀/필드 헬퍼
# ──────────────────────────────────────────────────────────────────────
def _cell_to_str(value: object) -> str | None:
    """셀 값을 문자열로. 정제 후 빈 값이면 None(미존재로 취급)."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _canon_revision(value: str) -> str:
    """교육과정 표기 정규화 — 미등록 표기는 원본 유지(모델 검증에 맡김)."""
    key = value.strip()
    return _REVISION_CANON.get(key, key)


# ──────────────────────────────────────────────────────────────────────
# 순수 행 파서 (openpyxl 불요 — transform 레이어 재사용)
# ──────────────────────────────────────────────────────────────────────
def parse_standard_row(row: Mapping[str, object]) -> AchievementStandard:
    """File A 성취기준 행(dict) → 검증된 `AchievementStandard`.

    `_STD_COLS`로 컬럼을 Raw 키에 매핑한 뒤 `transform_raw_to_standard`에 위임한다
    (norm_id 파생·학교급/학년군 추론·정제·Pydantic 검증 일체를 transform이 담당).
    """
    raw: dict[str, str] = {}
    for column, key in _STD_COLS.items():
        value = _cell_to_str(row.get(column))
        if value is None:
            continue
        raw[key] = _canon_revision(value) if key == "curriculum_revision" else value
    # source_url 의무 — 행에 없으면 NCIC 기본값(공공누리 1유형 출처 표시)
    raw.setdefault("source_url", _DEFAULT_SOURCE_URL)
    return transform_raw_to_standard(cast(RawStandardRecord, raw))


def parse_link_row(row: Mapping[str, object]) -> list[ConceptStandardLink]:
    """File A 연결 행(dict) → 검증된 `ConceptStandardLink` 목록(행→복수 가능).

    실 `개념-성취기준-CCSS` 시트 특성 3종을 처리한다:
      1. 교육과정 열이 없으면 revision은 **2022 개정 기본**(norm_id 직접값이 있으면 우선).
      2. link_type 괄호 접미사(`직접(기본수학)`·`재매핑(2015→2022)`)는 베어 토큰으로 정규화하고,
         괄호 내부는 `note`로 보존한다(비고 열이 따로 없을 때만 — 정보 손실 0).
      3. code 셀이 `;`로 복수 성취기준을 묶을 수 있어 **코드별 1 링크로 분할**한다.

    norm_id 직접값이 있으면 단일 링크로 쓰고, 없으면 code(복수 가능)에서 `build_norm_id` 파생한다.
    필수 필드 누락·link_type 비허용은 transform이 거른다.
    """
    mapped: dict[str, str] = {}
    for column, key in _LINK_COLS.items():
        value = _cell_to_str(row.get(column))
        if value is not None:
            mapped[key] = value

    # link_type 괄호 정규화 + 괄호 내부 note 보존(비고 미지정 시)
    link_type_raw = mapped.get("link_type", "")
    link_type = re.sub(r"\(.*?\)", "", link_type_raw).strip()
    note = mapped.get("note")
    if note is None:
        paren = re.search(r"\((.*?)\)", link_type_raw)
        if paren and paren.group(1).strip():
            note = paren.group(1).strip()

    concept_src_id = mapped.get("concept_src_id", "")

    def _build(norm_id: str) -> ConceptStandardLink:
        raw: dict[str, str] = {
            "concept_src_id": concept_src_id,
            "norm_id": norm_id,
            "link_type": link_type,
        }
        if note:
            raw["note"] = note
        return transform_raw_to_link(cast(RawLinkRecord, raw))

    # norm_id 직접값 우선(단일 링크)
    direct = mapped.get("norm_id")
    if direct:
        return [_build(direct)]

    # 없으면 code(복수 가능)에서 파생 — revision 기본 2022
    code_cell = mapped.get("code")
    if not code_cell:
        raise ExtractError("연결 행에 norm_id(또는 성취기준 코드)가 없습니다")
    revision = mapped.get("curriculum_revision")
    revision = _canon_revision(revision) if revision else "2022 개정"
    codes = [c.strip() for c in code_cell.split(";") if c.strip()]
    return [_build(build_norm_id(revision, code)) for code in codes]


# ──────────────────────────────────────────────────────────────────────
# openpyxl 셸 (지연 import)
# ──────────────────────────────────────────────────────────────────────
def _load_workbook(path: str | Path) -> Any:
    """openpyxl 워크북 로드(read-only·값만). openpyxl 미설치 시 친절한 오류."""
    try:
        import openpyxl
    except ModuleNotFoundError as exc:  # pragma: no cover - 설치 환경 의존
        raise ExtractError(
            "openpyxl 미설치 — `pip install -e '.[xlsx]'`로 설치 후 재시도하세요."
        ) from exc
    return openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)


def _iter_sheet_rows(workbook: Any, sheet_name: str) -> Iterator[dict[str, object]]:
    """워크시트를 헤더 기준 dict 행으로 순회. 첫 행을 헤더로 본다."""
    if sheet_name not in workbook.sheetnames:
        raise ExtractError(f"시트 없음: {sheet_name!r} (가용: {list(workbook.sheetnames)})")
    worksheet = workbook[sheet_name]
    row_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration:
        return
    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    for raw_row in row_iter:
        # 전부 빈 행은 skip(꼬리 공백 행 방어)
        if all(cell is None for cell in raw_row):
            continue
        record: dict[str, object] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = raw_row[index] if index < len(raw_row) else None
        yield record


def extract_file_a(
    path: str | Path,
) -> tuple[list[AchievementStandard], list[ConceptStandardLink]]:
    """File A xlsx → (성취기준 목록, 개념↔성취기준 연결 목록).

    `_SHEET_STANDARDS`·`_SHEET_LINKS` 시트를 읽어 각 행을 순수 파서로 변환한다. 변환·검증 실패는
    `TransformError`/`ExtractError`로 전파된다(조용히 넘기지 않는다 — CLAUDE.md 신뢰 원칙).
    """
    workbook = _load_workbook(path)
    try:
        standards = [
            parse_standard_row(row) for row in _iter_sheet_rows(workbook, _SHEET_STANDARDS)
        ]
        # 연결 행은 다중코드 분할로 행→복수 링크가 될 수 있어 평탄화(extend).
        links = [
            link for row in _iter_sheet_rows(workbook, _SHEET_LINKS) for link in parse_link_row(row)
        ]
    finally:
        workbook.close()
    return standards, links


__all__ = [
    "ExtractError",
    "extract_file_a",
    "parse_link_row",
    "parse_standard_row",
]
