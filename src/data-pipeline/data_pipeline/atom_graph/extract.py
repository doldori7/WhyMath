"""원자 백본 xlsx → 행 dict 추출(openpyxl 지연 import).

`extract_xlsx.py`(File A 추출) 선례 미러: openpyxl은 `[xlsx]` extra라 *지연 import*한다 —
순수 정형화 레이어(transform)는 openpyxl 없이도 import·테스트 가능해야 한다([postgres]·[neo4j]
드라이버 지연 import와 동형).

원본 xlsx 4시트 중 정형화에 쓰는 2시트만 읽는다:
  - `원자_통합마스터`(1,837행·31열) → 원자 행 dict
  - `선수엣지_통합`(3,220행·11열) → 엣지 행 dict
`안내`·`요약`은 자기보고 메타라 읽지 않는다(데이터카드 §4·§5에 이미 반영).
"""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from typing import Any, Final

# 정형화 대상 시트명(원본 헤더).
SHEET_ATOMS: Final[str] = "원자_통합마스터"
SHEET_EDGES: Final[str] = "선수엣지_통합"


class ExtractError(ValueError):
    """xlsx 추출 실패."""


def _load_workbook(path: str | Path) -> Any:
    """openpyxl 워크북 로드(read-only·값만). openpyxl 미설치 시 친절한 오류."""
    try:
        import openpyxl
    except ModuleNotFoundError as exc:  # pragma: no cover - 설치 환경 의존
        raise ExtractError(
            "openpyxl 미설치 — `pip install -e '.[xlsx]'`로 설치 후 재시도하세요."
        ) from exc
    return openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)


def _iter_sheet_rows(workbook: Any, sheet_name: str) -> Iterator[dict[str, object]]:
    """워크시트를 헤더 기준 dict 행으로 순회. 첫 행을 헤더로 본다. 전부 빈 행은 skip."""
    if sheet_name not in workbook.sheetnames:
        raise ExtractError(f"시트 없음: {sheet_name!r} (가용: {list(workbook.sheetnames)})")
    worksheet = workbook[sheet_name]
    row_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration:
        return
    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    for raw_row in row_iter:
        if all(cell is None for cell in raw_row):
            continue
        record: dict[str, object] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = raw_row[index] if index < len(raw_row) else None
        yield record


def extract_workbook(
    path: str | Path,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    """원자 백본 xlsx → (원자 행 목록, 엣지 행 목록).

    `SHEET_ATOMS`·`SHEET_EDGES` 시트를 읽어 행 dict로 변환한다(정형화는 transform 책임).
    """
    workbook = _load_workbook(path)
    try:
        atom_rows = list(_iter_sheet_rows(workbook, SHEET_ATOMS))
        edge_rows = list(_iter_sheet_rows(workbook, SHEET_EDGES))
    finally:
        workbook.close()
    return atom_rows, edge_rows


__all__ = [
    "SHEET_ATOMS",
    "SHEET_EDGES",
    "ExtractError",
    "extract_workbook",
]
