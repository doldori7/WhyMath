"""대학 성취기준 통합마스터 xlsx → 행 dict 추출 (openpyxl 지연 import).

`atom_graph/extract.py`·`ncic/extract_xlsx.py` 선례 미러: openpyxl은 `[xlsx]` extra라 *지연
import*한다(순수 정형화 레이어는 openpyxl 없이 import·테스트 가능).

업로드 파일("수학과 성취기준 통합 마스터")은 다중 시트(개념·성취기준_목록·노드_트리구조 등)지만,
대학 성취기준 정형화엔 **`성취기준_목록` 시트 한 장**만 읽는다 — 그 시트의 대학 행이 성취기준 본문과
소단원(개념ID) 링크를 모두 담는다(소단원↔성취기준 1:1). `개념` 시트의 콘텐츠(은유·오개념 등)는
별도 단계(U4) 몫이라 여기서 읽지 않는다.
"""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from typing import Any, Final

# 정형화 대상 시트명(업로드 파일 헤더).
SHEET_STANDARDS: Final[str] = "성취기준_목록"
# 대학 행 식별 컬럼·값(학교급=='대학교').
COL_SCHOOL_LEVEL: Final[str] = "학교급"
UNIVERSITY_SCHOOL_LEVEL: Final[str] = "대학교"


class ExtractError(ValueError):
    """xlsx 추출 실패."""


def _load_workbook(path: str | Path) -> Any:
    """openpyxl 워크북 로드(read-only·값만). openpyxl 미설치 시 친절한 오류(atom_graph 미러)."""
    try:
        import openpyxl
    except ModuleNotFoundError as exc:  # pragma: no cover - 설치 환경 의존
        raise ExtractError(
            "openpyxl 미설치 — `pip install -e '.[xlsx]'`로 설치 후 재시도하세요."
        ) from exc
    return openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)


def _iter_sheet_rows(workbook: Any, sheet_name: str) -> Iterator[dict[str, object]]:
    """워크시트를 헤더 기준 dict 행으로 순회(첫 행=헤더·전부 빈 행 skip·atom_graph 미러)."""
    if sheet_name not in workbook.sheetnames:
        raise ExtractError(f"시트 없음: {sheet_name!r} (가용: {list(workbook.sheetnames)})")
    worksheet = workbook[sheet_name]
    row_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration:
        return
    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    for raw_row in row_iter:
        if all(cell is None for cell in raw_row):
            continue
        record: dict[str, object] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = raw_row[index] if index < len(raw_row) else None
        yield record


def extract_university_standard_rows(path: str | Path) -> list[dict[str, object]]:
    """업로드 xlsx → 대학 성취기준 행 목록(`성취기준_목록`의 `학교급=='대학교'` 행만).

    정형화(transform)가 행 dict를 `UniversityStandard`/링크로 변환한다. 대학 외(초중고) 행은
    NCIC 공공누리 본문이라 이 자체작성 파이프라인에서 *읽지 않는다*(라이선스 분리·구조적 차단).
    """
    workbook = _load_workbook(path)
    try:
        rows = [
            row
            for row in _iter_sheet_rows(workbook, SHEET_STANDARDS)
            if str(row.get(COL_SCHOOL_LEVEL, "")).strip() == UNIVERSITY_SCHOOL_LEVEL
        ]
    finally:
        workbook.close()
    return rows


__all__ = [
    "COL_SCHOOL_LEVEL",
    "SHEET_STANDARDS",
    "UNIVERSITY_SCHOOL_LEVEL",
    "ExtractError",
    "extract_university_standard_rows",
]
